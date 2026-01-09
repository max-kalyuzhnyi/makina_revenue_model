"""Makina Revenue Model - Interactive Dashboard"""

import streamlit as st
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pandas as pd
import openpyxl
from copy import deepcopy

# Utility functions
def format_number(num):
    """Format numbers with M/B suffixes"""
    if num >= 1_000_000_000:
        return f"${num/1_000_000_000:.2f}B"
    elif num >= 1_000_000:
        return f"${num/1_000_000:.2f}M"
    elif num >= 1_000:
        return f"${num/1_000:.2f}K"
    else:
        return f"${num:.2f}"

def format_percentage(num):
    """Format percentage with 2 decimals"""
    return f"{num:.2f}%"

# Page configuration
st.set_page_config(
    page_title="Makina Revenue Model",
    page_icon="💰",
    layout="wide"
)

# Custom CSS
st.markdown("""
    <style>
        .block-container {
            padding-top: 2rem;
            padding-bottom: 0rem;
        }
        .stTabs [data-baseweb="tab-list"] {
            gap: 8px;
            padding-top: 1rem;
            padding-bottom: 1rem;
        }
        .stTabs [data-baseweb="tab"] {
            height: 50px;
            padding-left: 20px;
            padding-right: 20px;
            font-size: 16px;
        }
    </style>
    """, unsafe_allow_html=True)

# Data loading function
@st.cache_data
def load_excel_data(file_path):
    """Load data from Excel file"""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    data = {}
    for year in [2025, 2026, 2027]:
        sheet_name = f'MAK Revenue {year}'
        sheet = wb[sheet_name]

        # Extract inputs from modeling section
        # Row 9: TVL in native units (input)
        # Row 10: Prices (input)
        # Row 12-14: Fee rates and performance growth (input)
        year_data = {
            'year': year,
            'usdc': {
                'tvl_units': sheet.cell(9, 3).value or 0,  # Native units
                'price': sheet.cell(10, 3).value or 1.0,
                'mgmt_fee': sheet.cell(12, 3).value or 0,
                'perf_fee': sheet.cell(13, 3).value or 0,
                'performance_growth': sheet.cell(14, 3).value or 0,
            },
            'eth': {
                'tvl_units': sheet.cell(9, 7).value or 0,  # Native units (ETH)
                'price': sheet.cell(10, 7).value or 3000.0,
                'mgmt_fee': sheet.cell(12, 7).value or 0,
                'perf_fee': sheet.cell(13, 7).value or 0,
                'performance_growth': sheet.cell(14, 7).value or 0,
            },
            'btc': {
                'tvl_units': sheet.cell(9, 11).value or 0,  # Native units (BTC)
                'price': sheet.cell(10, 11).value or 90000.0,
                'mgmt_fee': sheet.cell(12, 11).value or 0,
                'perf_fee': sheet.cell(13, 11).value or 0,
                'performance_growth': sheet.cell(14, 11).value or 0,
            },
            'fee_split': {
                'mgmt_dao': sheet.cell(17, 3).value or 0.6,
                'mgmt_operator': sheet.cell(17, 4).value or 0.4,
                'perf_dao': sheet.cell(18, 3).value or 0.6,
                'perf_operator': sheet.cell(18, 4).value or 0.4,
            },
            'valuation': {
                'price_rev_ratio': sheet.cell(4, 8).value or 45.0,
            },
            'dao_split': {
                'operations': sheet.cell(4, 11).value or 0.3,
                'buyback': sheet.cell(5, 11).value or 0.4,
                'revenue_share': sheet.cell(6, 11).value or 0.3,
            }
        }
        data[year] = year_data

    return data

# Calculation functions
def calculate_revenue(data):
    """Calculate revenue for each asset and total"""
    results = {}

    for asset in ['usdc', 'eth', 'btc']:
        asset_data = data[asset]

        # Calculate TVL in USD (native units × price)
        tvl_usd = asset_data['tvl_units'] * asset_data['price']

        # Management fee revenue (in native currency)
        mgmt_revenue_native = asset_data['tvl_units'] * asset_data['mgmt_fee']
        mgmt_revenue_usd = mgmt_revenue_native * asset_data['price']

        # Performance fee revenue (in native currency)
        perf_revenue_native = asset_data['tvl_units'] * asset_data['performance_growth'] * asset_data['perf_fee']
        perf_revenue_usd = perf_revenue_native * asset_data['price']

        # Total revenue
        total_revenue_native = mgmt_revenue_native + perf_revenue_native
        total_revenue_usd = mgmt_revenue_usd + perf_revenue_usd

        results[asset] = {
            'tvl_usd': tvl_usd,
            'mgmt_revenue_native': mgmt_revenue_native,
            'mgmt_revenue_usd': mgmt_revenue_usd,
            'perf_revenue_native': perf_revenue_native,
            'perf_revenue_usd': perf_revenue_usd,
            'total_revenue_native': total_revenue_native,
            'total_revenue_usd': total_revenue_usd,
        }

    # Calculate splits by DAO/Operator
    fee_split = data['fee_split']

    # Total management revenue (in USD)
    total_mgmt = sum(results[asset]['mgmt_revenue_usd'] for asset in ['usdc', 'eth', 'btc'])
    mgmt_dao = total_mgmt * fee_split['mgmt_dao']
    mgmt_operator = total_mgmt * fee_split['mgmt_operator']

    # Total performance revenue (in USD)
    total_perf = sum(results[asset]['perf_revenue_usd'] for asset in ['usdc', 'eth', 'btc'])
    perf_dao = total_perf * fee_split['perf_dao']
    perf_operator = total_perf * fee_split['perf_operator']

    # Totals
    total_tvl = sum(results[asset]['tvl_usd'] for asset in ['usdc', 'eth', 'btc'])
    total_revenue = total_mgmt + total_perf
    dao_revenue = mgmt_dao + perf_dao
    operator_revenue = mgmt_operator + perf_operator

    # Calculate take rates
    dao_take_rate = (dao_revenue / total_tvl * 100) if total_tvl > 0 else 0
    total_take_rate = (total_revenue / total_tvl * 100) if total_tvl > 0 else 0

    # Valuation metrics
    price_rev_ratio = data['valuation']['price_rev_ratio']
    fdv = dao_revenue * price_rev_ratio
    fdv_tvl_ratio = (fdv / total_tvl) if total_tvl > 0 else 0

    results['summary'] = {
        'total_tvl': total_tvl,
        'total_revenue': total_revenue,
        'dao_revenue': dao_revenue,
        'operator_revenue': operator_revenue,
        'mgmt_dao': mgmt_dao,
        'mgmt_operator': mgmt_operator,
        'perf_dao': perf_dao,
        'perf_operator': perf_operator,
        'dao_take_rate': dao_take_rate,
        'total_take_rate': total_take_rate,
        'price_rev_ratio': price_rev_ratio,
        'fdv': fdv,
        'fdv_tvl_ratio': fdv_tvl_ratio,
    }

    return results

# Initialize session state with Excel data
if 'data' not in st.session_state:
    excel_file = 'Makina Revenue Generation Estimates.xlsx'
    st.session_state.data = load_excel_data(excel_file)
    st.session_state.original_data = deepcopy(st.session_state.data)

# Sidebar - title
with st.sidebar:
    st.markdown("**💰 Makina Revenue Model**")
    st.markdown("---")
    st.markdown("Navigate between tabs to view dashboard and configure parameters")

# Main content
tab1, tab2 = st.tabs(["📊 Dashboard", "⚙️ Configuration"])

# TAB 1: Dashboard
with tab1:
    st.header("Revenue Model Dashboard")
    st.markdown("### End of 2027 Summary")

    # Calculate results for all years
    all_results = {}
    for year in [2025, 2026, 2027]:
        all_results[year] = calculate_revenue(st.session_state.data[year])

    # Get 2027 summary for top metrics
    summary_2027 = all_results[2027]['summary']

    # Top metrics row
    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        st.metric("Total TVL", format_number(summary_2027['total_tvl']))
    with col2:
        st.metric("Total Revenue", format_number(summary_2027['total_revenue']))
    with col3:
        st.metric("DAO Revenue", format_number(summary_2027['dao_revenue']))
    with col4:
        st.metric("DAO Take Rate", format_percentage(summary_2027['dao_take_rate']))
    with col5:
        # Add P/Rev ratio control
        price_rev = st.number_input(
            "P/Rev Ratio",
            value=summary_2027['price_rev_ratio'],
            min_value=0.0,
            step=1.0,
            key='price_rev_dashboard'
        )
        # Update all years if changed
        if price_rev != summary_2027['price_rev_ratio']:
            for year in [2025, 2026, 2027]:
                st.session_state.data[year]['valuation']['price_rev_ratio'] = price_rev
            st.rerun()

    # Second row with FDV metrics
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("FDV", format_number(summary_2027['fdv']))
    with col2:
        st.metric("FDV/TVL Ratio", f"{summary_2027['fdv_tvl_ratio']:.2f}x")

    st.markdown("---")

    # Prepare data for all charts
    tvl_data = []
    for year in [2025, 2026, 2027]:
        results = all_results[year]
        tvl_data.append({
            'Year': year,
            'USDC': results['usdc']['tvl_usd'],
            'ETH': results['eth']['tvl_usd'],
            'BTC': results['btc']['tvl_usd'],
        })
    tvl_df = pd.DataFrame(tvl_data)

    # Define colors
    colors = {
        'USDC': '#2ecc71',  # Green
        'ETH': '#3498db',   # Blue
        'BTC': '#f39c12'    # Orange
    }

    mgmt_fees = [all_results[year]['summary']['mgmt_dao'] for year in [2025, 2026, 2027]]
    perf_fees = [all_results[year]['summary']['perf_dao'] for year in [2025, 2026, 2027]]
    fdv_values = [all_results[year]['summary']['fdv'] for year in [2025, 2026, 2027]]
    fdv_tvl_values = [all_results[year]['summary']['fdv_tvl_ratio'] for year in [2025, 2026, 2027]]
    take_rates = [all_results[year]['summary']['dao_take_rate'] for year in [2025, 2026, 2027]]

    # First row of charts: TVL by Asset and DAO Fees
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("TVL by Asset (2025-2027)")
        fig_tvl = go.Figure()

        for asset in ['BTC', 'ETH', 'USDC']:
            fig_tvl.add_trace(go.Bar(
                x=tvl_df['Year'],
                y=tvl_df[asset],
                name=asset,
                marker_color=colors[asset],
                hovertemplate='%{y:,.0f}<extra></extra>'
            ))

        fig_tvl.update_layout(
            barmode='stack',
            xaxis_title="Year",
            yaxis_title="TVL (USD)",
            height=400,
            yaxis=dict(tickformat=',.0f'),
            hovermode='x unified',
            margin=dict(l=20, r=20, t=20, b=20)
        )
        st.plotly_chart(fig_tvl, use_container_width=True)

    with col2:
        st.subheader("DAO Fees by Type (2025-2027)")
        fig_fees = go.Figure()

        fig_fees.add_trace(go.Bar(
            x=[2025, 2026, 2027],
            y=mgmt_fees,
            name='Management Fees',
            marker_color='#2ecc71',
            hovertemplate='%{y:,.0f}<extra></extra>'
        ))

        fig_fees.add_trace(go.Bar(
            x=[2025, 2026, 2027],
            y=perf_fees,
            name='Performance Fees',
            marker_color='#3498db',
            hovertemplate='%{y:,.0f}<extra></extra>'
        ))

        fig_fees.update_layout(
            barmode='stack',
            xaxis_title="Year",
            yaxis_title="DAO Fees (USD)",
            height=400,
            yaxis=dict(tickformat=',.0f'),
            hovermode='x unified',
            margin=dict(l=20, r=20, t=20, b=20)
        )
        st.plotly_chart(fig_fees, use_container_width=True)

    st.markdown("---")

    # Second row of charts: FDV & FDV/TVL and DAO Take Rate
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("FDV and FDV/TVL Ratio (2025-2027)")
        fig_fdv = make_subplots(specs=[[{"secondary_y": True}]])

        # Add bar chart for FDV
        fig_fdv.add_trace(
            go.Bar(
                x=[2025, 2026, 2027],
                y=fdv_values,
                name="FDV",
                marker_color='#9b59b6',
                hovertemplate='%{y:,.0f}<extra></extra>'
            ),
            secondary_y=False
        )

        # Add line chart for FDV/TVL ratio
        fig_fdv.add_trace(
            go.Scatter(
                x=[2025, 2026, 2027],
                y=fdv_tvl_values,
                name="FDV/TVL Ratio",
                mode='lines+markers',
                line=dict(color='#e74c3c', width=3),
                marker=dict(size=10),
                hovertemplate='%{y:.2f}x<extra></extra>'
            ),
            secondary_y=True
        )

        fig_fdv.update_xaxes(title_text="Year")
        fig_fdv.update_yaxes(title_text="FDV (USD)", secondary_y=False, tickformat=',.0f')
        fig_fdv.update_yaxes(title_text="FDV/TVL Ratio", secondary_y=True)
        fig_fdv.update_layout(
            height=400,
            hovermode='x unified',
            margin=dict(l=20, r=20, t=20, b=20)
        )
        st.plotly_chart(fig_fdv, use_container_width=True)

    with col2:
        st.subheader("DAO Take Rate (2025-2027)")
        fig_take = go.Figure()

        fig_take.add_trace(go.Scatter(
            x=[2025, 2026, 2027],
            y=take_rates,
            mode='lines+markers',
            line=dict(color='#e74c3c', width=3),
            marker=dict(size=12),
            hovertemplate='%{y:.2f}%<extra></extra>'
        ))

        fig_take.update_layout(
            xaxis_title="Year",
            yaxis_title="DAO Take Rate (%)",
            height=400,
            yaxis=dict(tickformat='.2f', ticksuffix='%'),
            hovermode='x unified',
            margin=dict(l=20, r=20, t=20, b=20)
        )
        st.plotly_chart(fig_take, use_container_width=True)

# TAB 2: Configuration
with tab2:
    st.header("Configuration")

    # Year selector
    selected_year = st.selectbox("Select Year", [2025, 2026, 2027], key='config_year')

    year_data = st.session_state.data[selected_year]

    # Collect all inputs first
    st.markdown("### Asset Parameters")

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("**USDC**")
        usdc_tvl = st.number_input("TVL Amount (USDC)", value=float(year_data['usdc']['tvl_units']),
                                   min_value=0.0, step=1000000.0, format="%.0f", key=f'usdc_tvl_{selected_year}')
        usdc_price = st.number_input("Price (USD)", value=year_data['usdc']['price'],
                                     min_value=0.0, step=0.01, key=f'usdc_price_{selected_year}')
        usdc_mgmt = st.number_input("Mgmt Fee (%)", value=year_data['usdc']['mgmt_fee']*100,
                                    min_value=0.0, max_value=100.0, step=0.1, format="%.2f", key=f'usdc_mgmt_{selected_year}')
        usdc_perf = st.number_input("Perf Fee (%)", value=year_data['usdc']['perf_fee']*100,
                                    min_value=0.0, max_value=100.0, step=0.1, format="%.2f", key=f'usdc_perf_{selected_year}')
        usdc_growth = st.number_input("Performance Growth (%)", value=year_data['usdc']['performance_growth']*100,
                                      min_value=0.0, max_value=100.0, step=0.1, format="%.2f", key=f'usdc_growth_{selected_year}')

    with col2:
        st.markdown("**ETH**")
        eth_tvl = st.number_input("TVL Amount (ETH)", value=float(year_data['eth']['tvl_units']),
                                 min_value=0.0, step=1000.0, format="%.0f", key=f'eth_tvl_{selected_year}')
        eth_price = st.number_input("Price (USD)", value=year_data['eth']['price'],
                                    min_value=0.0, step=100.0, key=f'eth_price_{selected_year}')
        eth_mgmt = st.number_input("Mgmt Fee (%)", value=year_data['eth']['mgmt_fee']*100,
                                   min_value=0.0, max_value=100.0, step=0.1, format="%.2f", key=f'eth_mgmt_{selected_year}')
        eth_perf = st.number_input("Perf Fee (%)", value=year_data['eth']['perf_fee']*100,
                                   min_value=0.0, max_value=100.0, step=0.1, format="%.2f", key=f'eth_perf_{selected_year}')
        eth_growth = st.number_input("Performance Growth (%)", value=year_data['eth']['performance_growth']*100,
                                     min_value=0.0, max_value=100.0, step=0.1, format="%.2f", key=f'eth_growth_{selected_year}')

    with col3:
        st.markdown("**BTC**")
        btc_tvl = st.number_input("TVL Amount (BTC)", value=float(year_data['btc']['tvl_units']),
                                 min_value=0.0, step=10.0, format="%.0f", key=f'btc_tvl_{selected_year}')
        btc_price = st.number_input("Price (USD)", value=year_data['btc']['price'],
                                    min_value=0.0, step=1000.0, key=f'btc_price_{selected_year}')
        btc_mgmt = st.number_input("Mgmt Fee (%)", value=year_data['btc']['mgmt_fee']*100,
                                   min_value=0.0, max_value=100.0, step=0.1, format="%.2f", key=f'btc_mgmt_{selected_year}')
        btc_perf = st.number_input("Perf Fee (%)", value=year_data['btc']['perf_fee']*100,
                                   min_value=0.0, max_value=100.0, step=0.1, format="%.2f", key=f'btc_perf_{selected_year}')
        btc_growth = st.number_input("Performance Growth (%)", value=year_data['btc']['performance_growth']*100,
                                     min_value=0.0, max_value=100.0, step=0.1, format="%.2f", key=f'btc_growth_{selected_year}')

    st.markdown("---")
    st.markdown("### DAO-Related Parameters")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**Fee Split (DAO / Operator)**")
        mgmt_dao_split = st.number_input("Management Fee - DAO Share (%)", value=year_data['fee_split']['mgmt_dao']*100,
                                         min_value=0.0, max_value=100.0, step=1.0, key=f'mgmt_dao_{selected_year}')
        mgmt_op_split = 100 - mgmt_dao_split
        st.text(f"Management Fee - Operator: {mgmt_op_split:.1f}%")

        st.markdown("")
        perf_dao_split = st.number_input("Performance Fee - DAO Share (%)", value=year_data['fee_split']['perf_dao']*100,
                                         min_value=0.0, max_value=100.0, step=1.0, key=f'perf_dao_{selected_year}')
        perf_op_split = 100 - perf_dao_split
        st.text(f"Performance Fee - Operator: {perf_op_split:.1f}%")

    with col2:
        st.markdown("**DAO Revenue Allocation**")
        dao_ops = st.number_input("DAO Operations (%)", value=year_data['dao_split']['operations']*100,
                                  min_value=0.0, max_value=100.0, step=1.0, key=f'dao_ops_{selected_year}')
        dao_buyback = st.number_input("Buyback (%)", value=year_data['dao_split']['buyback']*100,
                                      min_value=0.0, max_value=100.0, step=1.0, key=f'dao_buyback_{selected_year}')
        dao_revshare = 100 - dao_ops - dao_buyback
        st.text(f"Revenue Share: {dao_revshare:.1f}%")

        if dao_revshare < 0:
            st.warning(f"⚠️ Operations + Buyback cannot exceed 100%")

    st.markdown("---")
    st.markdown("### Valuation Parameters")

    col1, col2 = st.columns(2)
    with col1:
        price_rev_ratio = st.number_input("Price/Revenue Ratio", value=year_data['valuation']['price_rev_ratio'],
                                         min_value=0.0, step=1.0, key=f'price_rev_{selected_year}')

    # Now calculate live preview and show buttons at the top
    st.markdown("---")
    st.markdown("---")  # Extra separator

    # Action buttons FIRST
    col1, col2, col3 = st.columns([1, 1, 3])

    with col1:
        save_button = st.button("💾 Save Changes", type="primary", key=f'save_{selected_year}')
    with col2:
        reset_button = st.button("🔄 Reset to Original", key=f'reset_{selected_year}')

    st.markdown("---")

    # Calculate live preview based on current inputs
    st.subheader(f"📊 Calculated Results (Live Preview)")

    # Build temporary data structure with current input values
    temp_data = {
        'usdc': {
            'tvl_units': usdc_tvl,
            'price': usdc_price,
            'mgmt_fee': usdc_mgmt / 100,
            'perf_fee': usdc_perf / 100,
            'performance_growth': usdc_growth / 100,
        },
        'eth': {
            'tvl_units': eth_tvl,
            'price': eth_price,
            'mgmt_fee': eth_mgmt / 100,
            'perf_fee': eth_perf / 100,
            'performance_growth': eth_growth / 100,
        },
        'btc': {
            'tvl_units': btc_tvl,
            'price': btc_price,
            'mgmt_fee': btc_mgmt / 100,
            'perf_fee': btc_perf / 100,
            'performance_growth': btc_growth / 100,
        },
        'fee_split': {
            'mgmt_dao': mgmt_dao_split / 100,
            'mgmt_operator': (100 - mgmt_dao_split) / 100,
            'perf_dao': perf_dao_split / 100,
            'perf_operator': (100 - perf_dao_split) / 100,
        },
        'valuation': {
            'price_rev_ratio': price_rev_ratio,
        },
        'dao_split': {
            'operations': dao_ops / 100,
            'buyback': dao_buyback / 100,
            'revenue_share': (100 - dao_ops - dao_buyback) / 100,
        }
    }

    # Calculate with live values
    live_results = calculate_revenue(temp_data)
    live_summary = live_results['summary']

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("Total TVL", format_number(live_summary['total_tvl']))
        st.metric("USDC TVL", format_number(live_results['usdc']['tvl_usd']))

    with col2:
        st.metric("Total Revenue", format_number(live_summary['total_revenue']))
        st.metric("ETH TVL", format_number(live_results['eth']['tvl_usd']))

    with col3:
        st.metric("DAO Revenue", format_number(live_summary['dao_revenue']))
        st.metric("BTC TVL", format_number(live_results['btc']['tvl_usd']))

    with col4:
        st.metric("DAO Take Rate", format_percentage(live_summary['dao_take_rate']))
        st.metric("FDV", format_number(live_summary['fdv']))

    # Handle button actions
    if save_button:
        # Update session state with new values
        st.session_state.data[selected_year]['usdc']['tvl_units'] = usdc_tvl
        st.session_state.data[selected_year]['usdc']['price'] = usdc_price
        st.session_state.data[selected_year]['usdc']['mgmt_fee'] = usdc_mgmt / 100
        st.session_state.data[selected_year]['usdc']['perf_fee'] = usdc_perf / 100
        st.session_state.data[selected_year]['usdc']['performance_growth'] = usdc_growth / 100

        st.session_state.data[selected_year]['eth']['tvl_units'] = eth_tvl
        st.session_state.data[selected_year]['eth']['price'] = eth_price
        st.session_state.data[selected_year]['eth']['mgmt_fee'] = eth_mgmt / 100
        st.session_state.data[selected_year]['eth']['perf_fee'] = eth_perf / 100
        st.session_state.data[selected_year]['eth']['performance_growth'] = eth_growth / 100

        st.session_state.data[selected_year]['btc']['tvl_units'] = btc_tvl
        st.session_state.data[selected_year]['btc']['price'] = btc_price
        st.session_state.data[selected_year]['btc']['mgmt_fee'] = btc_mgmt / 100
        st.session_state.data[selected_year]['btc']['perf_fee'] = btc_perf / 100
        st.session_state.data[selected_year]['btc']['performance_growth'] = btc_growth / 100

        st.session_state.data[selected_year]['fee_split']['mgmt_dao'] = mgmt_dao_split / 100
        st.session_state.data[selected_year]['fee_split']['mgmt_operator'] = (100 - mgmt_dao_split) / 100
        st.session_state.data[selected_year]['fee_split']['perf_dao'] = perf_dao_split / 100
        st.session_state.data[selected_year]['fee_split']['perf_operator'] = (100 - perf_dao_split) / 100

        st.session_state.data[selected_year]['valuation']['price_rev_ratio'] = price_rev_ratio

        st.session_state.data[selected_year]['dao_split']['operations'] = dao_ops / 100
        st.session_state.data[selected_year]['dao_split']['buyback'] = dao_buyback / 100
        st.session_state.data[selected_year]['dao_split']['revenue_share'] = (100 - dao_ops - dao_buyback) / 100

        st.success(f"✅ Changes saved for {selected_year}!")
        st.rerun()

    if reset_button:
        st.session_state.data[selected_year] = deepcopy(st.session_state.original_data[selected_year])
        st.success(f"✅ Reset {selected_year} to original values!")
        st.rerun()
