"""Makina Revenue Model - Simplified Dashboard"""

import streamlit as st
import plotly.graph_objects as go
import pandas as pd
import openpyxl
from copy import deepcopy

# Utility functions
def format_number(num):
    """Format numbers with M/B suffixes"""
    if num >= 1_000_000_000:
        return f"${num/1_000_000_000:.2f}B"
    elif num >= 1_000_000:
        return f"${num/1_000_000:.2f}M"
    elif num >= 1_000:
        return f"${num/1_000:.2f}K"
    else:
        return f"${num:.2f}"

def format_number_short(num):
    """Format numbers with M/B suffixes (compact - no decimals for M, 1 decimal for B)"""
    if num >= 1_000_000_000:
        return f"${num/1_000_000_000:.1f}B"
    elif num >= 1_000_000:
        return f"${num/1_000_000:.0f}M"
    elif num >= 1_000:
        return f"${num/1_000:.0f}K"
    else:
        return f"${num:.0f}"

def format_percentage(num):
    """Format percentage with 2 decimals"""
    return f"{num:.2f}%"

# Page configuration
st.set_page_config(
    page_title="Makina Revenue Model",
    page_icon="💰",
    layout="wide"
)

# Custom CSS
st.markdown("""
    <style>
        .block-container {
            padding-top: 2rem;
            padding-bottom: 0rem;
        }

        /* Make buttons blue instead of red */
        .stButton > button {
            background-color: #0068C9;
            color: white;
        }
        .stButton > button:hover {
            background-color: #0054A3;
            color: white;
        }
        .stButton > button:active {
            background-color: #003D75;
        }

        /* Make sliders blue instead of red */
        /* Slider thumb (the draggable circle) */
        div[data-baseweb="slider"] div[role="slider"] {
            background-color: #0068C9 !important;
        }
        /* Slider track (the filled portion) */
        div[data-baseweb="slider"] div[data-testid="stThumbValue"] {
            color: #0068C9 !important;
        }
    </style>
    <script>
    // Format slider numbers with thousand separators
    function formatSliderNumbers() {
        const sliders = document.querySelectorAll('[data-testid="stSlider"]');
        sliders.forEach(slider => {
            const thumbValue = slider.querySelector('[data-testid="stThumbValue"]');
            if (thumbValue && thumbValue.textContent) {
                const num = parseFloat(thumbValue.textContent.replace(/,/g, ''));
                if (!isNaN(num) && num >= 1000) {
                    thumbValue.textContent = num.toLocaleString('en-US');
                }
            }
        });
    }

    // Run on load and on slider changes
    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', formatSliderNumbers);
    } else {
        formatSliderNumbers();
    }

    // Watch for changes
    const observer = new MutationObserver(formatSliderNumbers);
    observer.observe(document.body, { childList: true, subtree: true });
    </script>
    """, unsafe_allow_html=True)

# Data loading function
@st.cache_data
def load_excel_data(file_path):
    """Load data from Excel file with custom starting values"""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    data = {}

    # 2025 specific values
    data[2025] = {
        'year': 2025,
        'usdc': {
            'tvl_units': 55_000_000,  # 55M
            'price': 1.0,
            'mgmt_fee': 0.01,  # 1%
            'perf_fee': 0.15,  # 15%
            'performance_growth': 0.04,  # 4% APR
        },
        'eth': {
            'tvl_units': 9_300,  # 9,300 ETH
            'price': 3_000.0,
            'mgmt_fee': 0.0075,  # 0.75%
            'perf_fee': 0.13,  # 13%
            'performance_growth': 0.02,  # 2% APR
        },
        'btc': {
            'tvl_units': 200,  # 200 BTC
            'price': 90_000.0,
            'mgmt_fee': 0.005,  # 0.5%
            'perf_fee': 0.10,  # 10%
            'performance_growth': 0.016,  # 1.6% APR
        },
        'fee_split': {
            'mgmt_dao': 0.4,  # 40% DAO
            'mgmt_operator': 0.6,  # 60% Operator
            'perf_dao': 0.4,  # 40% DAO
            'perf_operator': 0.6,  # 60% Operator
        },
        'valuation': {
            'price_rev_ratio': 45.0,
        },
        'buyback_pct': 70.0,
    }

    # 2026 and 2027 - read from Excel but override management fees and APRs
    for year in [2026, 2027]:
        sheet_name = f'MAK Revenue {year}'
        sheet = wb[sheet_name]

        year_data = {
            'year': year,
            'usdc': {
                'tvl_units': sheet.cell(9, 3).value or 0,
                'price': sheet.cell(10, 3).value or 1.0,
                'mgmt_fee': 0.01,  # Override to 1%
                'perf_fee': sheet.cell(13, 3).value or 0,
                'performance_growth': 0.10,  # Override to 10% APR
            },
            'eth': {
                'tvl_units': sheet.cell(9, 7).value or 0,
                'price': sheet.cell(10, 7).value or 3000.0,
                'mgmt_fee': 0.01,  # Override to 1%
                'perf_fee': sheet.cell(13, 7).value or 0,
                'performance_growth': 0.06,  # Override to 6% APR
            },
            'btc': {
                'tvl_units': sheet.cell(9, 11).value or 0,
                'price': sheet.cell(10, 11).value or 90000.0,
                'mgmt_fee': 0.01,  # Override to 1%
                'perf_fee': sheet.cell(13, 11).value or 0,
                'performance_growth': 0.03,  # Override to 3% APR
            },
            'fee_split': {
                'mgmt_dao': sheet.cell(17, 3).value or 0.6,
                'mgmt_operator': sheet.cell(17, 4).value or 0.4,
                'perf_dao': sheet.cell(18, 3).value or 0.6,
                'perf_operator': sheet.cell(18, 4).value or 0.4,
            },
            'valuation': {
                'price_rev_ratio': sheet.cell(4, 8).value or 45.0,
            },
            'buyback_pct': 70.0,
        }
        data[year] = year_data

    return data

# Calculate metrics for single asset
def calculate_asset_metrics(asset_data, fee_split):
    """Calculate metrics for a single asset"""
    # TVL in USD
    tvl_usd = asset_data['tvl_units'] * asset_data['price']

    # Total APR
    total_apr = asset_data['performance_growth']

    # APR breakdown
    mgmt_fee_apr = asset_data['mgmt_fee']
    perf_fee_apr = asset_data['performance_growth'] * asset_data['perf_fee']
    lp_apr = total_apr - mgmt_fee_apr - perf_fee_apr

    # Revenue calculations (annual)
    mgmt_revenue = tvl_usd * mgmt_fee_apr
    perf_revenue = tvl_usd * perf_fee_apr
    total_revenue = mgmt_revenue + perf_revenue

    # DAO split
    dao_mgmt = mgmt_revenue * fee_split['mgmt_dao']
    dao_perf = perf_revenue * fee_split['perf_dao']
    dao_revenue = dao_mgmt + dao_perf

    # Operator split
    op_mgmt = mgmt_revenue * fee_split['mgmt_operator']
    op_perf = perf_revenue * fee_split['perf_operator']
    op_revenue = op_mgmt + op_perf

    # DAO take rate
    dao_take_rate = (dao_revenue / tvl_usd * 100) if tvl_usd > 0 else 0

    # Operator take rate
    op_take_rate = (op_revenue / tvl_usd * 100) if tvl_usd > 0 else 0

    return {
        'tvl_usd': tvl_usd,
        'total_apr': total_apr * 100,  # Convert to percentage
        'mgmt_fee_apr': mgmt_fee_apr * 100,
        'perf_fee_apr': perf_fee_apr * 100,
        'lp_apr': lp_apr * 100,
        'total_revenue': total_revenue,
        'dao_revenue': dao_revenue,
        'op_revenue': op_revenue,
        'dao_take_rate': dao_take_rate,
        'op_take_rate': op_take_rate,
    }

# Create Sankey diagram for APR flow
def create_flow_chart(metrics, asset_name, dao_perf_share=0.6, dao_mgmt_share=0.6):
    """Create a Sankey diagram showing APR flow"""

    # Nodes
    labels = [
        "Total APR",           # 0
        "To LPs",              # 1
        "Performance Fee",     # 2
        "Management Fee",      # 3
        "DAO",                 # 4
        "Operator"             # 5
    ]

    # Links (source, target, value)
    total = metrics['total_apr']
    lp_pct = metrics['lp_apr']
    perf_pct = metrics['perf_fee_apr']
    mgmt_pct = metrics['mgmt_fee_apr']

    # From Total APR
    sources = [0, 0, 0]
    targets = [1, 2, 3]
    values = [lp_pct, perf_pct, mgmt_pct]

    # From Performance Fee to DAO/Operator (using actual split)
    if perf_pct > 0:
        sources.extend([2, 2])
        targets.extend([4, 5])
        values.extend([perf_pct * dao_perf_share, perf_pct * (1 - dao_perf_share)])

    # From Management Fee to DAO/Operator (using actual split)
    if mgmt_pct > 0:
        sources.extend([3, 3])
        targets.extend([4, 5])
        values.extend([mgmt_pct * dao_mgmt_share, mgmt_pct * (1 - dao_mgmt_share)])

    # Colors
    node_colors = ['#95a5a6', '#2ecc71', '#3498db', '#9b59b6', '#e67e22', '#e74c3c']

    fig = go.Figure(data=[go.Sankey(
        node=dict(
            pad=15,
            thickness=20,
            line=dict(color="white", width=0),
            label=labels,
            color=node_colors
        ),
        link=dict(
            source=sources,
            target=targets,
            value=values,
            color='rgba(200, 200, 200, 0.4)'
        ),
        textfont=dict(
            color='black',
            size=12,
            family='Arial, sans-serif'
        )
    )])

    fig.update_layout(
        title=dict(
            text=f"{asset_name} APR Flow (%)",
            font=dict(size=14, family="Arial, sans-serif", color="black")
        ),
        font=dict(size=12, family="Arial, sans-serif", color="black"),
        height=300,
        margin=dict(l=10, r=10, t=40, b=10),
        paper_bgcolor='white',
        plot_bgcolor='white'
    )

    return fig

# Initialize session state with Excel data
if 'data' not in st.session_state:
    excel_file = 'Makina Revenue Generation Estimates.xlsx'
    st.session_state.data = load_excel_data(excel_file)

# Initialize changes tracking
if 'has_changes' not in st.session_state:
    st.session_state.has_changes = False

# Track if data has been modified from original (for Reset button)
if 'data_modified' not in st.session_state:
    st.session_state.data_modified = False

# Main Dashboard
st.title("💰 Makina Revenue Model")

# Year selector and buttons at top
if st.session_state.has_changes:
    col_year, col_apply, col_reset = st.columns([2, 1, 1])
else:
    col_year, col_reset = st.columns([3, 1])

with col_year:
    selected_year = st.selectbox("Select Year", [2026, 2027], index=0, key='year_selector')

if st.session_state.has_changes:
    with col_apply:
        st.markdown("<div style='margin-top: 1.7rem;'></div>", unsafe_allow_html=True)
        if st.button("✓ Apply Changes", type="primary", use_container_width=True):
            # Apply buyback_pct and price_rev
            buyback_key = f'buyback_slider_{selected_year}'
            price_rev_key = f'price_rev_slider_{selected_year}'

            if buyback_key in st.session_state:
                st.session_state.data[selected_year]['buyback_pct'] = st.session_state[buyback_key]
            if price_rev_key in st.session_state:
                st.session_state.data[selected_year]['valuation']['price_rev_ratio'] = st.session_state[price_rev_key]

            # Apply all pending changes from slider values
            for asset_key in ['usdc', 'eth', 'btc']:
                # Get slider keys for this asset
                apr_key = f'{asset_key}_apr_{selected_year}'
                tvl_key = f'{asset_key}_tvl_{selected_year}'
                price_key = f'{asset_key}_price_{selected_year}'
                perf_fee_key = f'{asset_key}_perf_fee_{selected_year}'
                mgmt_fee_key = f'{asset_key}_mgmt_fee_{selected_year}'
                dao_perf_key = f'{asset_key}_dao_perf_{selected_year}'
                dao_mgmt_key = f'{asset_key}_dao_mgmt_{selected_year}'

                # Apply values if keys exist in session state
                if apr_key in st.session_state:
                    st.session_state.data[selected_year][asset_key]['performance_growth'] = st.session_state[apr_key] / 100
                if tvl_key in st.session_state:
                    st.session_state.data[selected_year][asset_key]['tvl_units'] = st.session_state[tvl_key]
                if price_key in st.session_state:
                    st.session_state.data[selected_year][asset_key]['price'] = st.session_state[price_key]
                if perf_fee_key in st.session_state:
                    st.session_state.data[selected_year][asset_key]['perf_fee'] = st.session_state[perf_fee_key] / 100
                if mgmt_fee_key in st.session_state:
                    st.session_state.data[selected_year][asset_key]['mgmt_fee'] = st.session_state[mgmt_fee_key] / 100
                if dao_perf_key in st.session_state:
                    st.session_state.data[selected_year]['fee_split']['perf_dao'] = st.session_state[dao_perf_key] / 100
                    st.session_state.data[selected_year]['fee_split']['perf_operator'] = (100 - st.session_state[dao_perf_key]) / 100
                if dao_mgmt_key in st.session_state:
                    st.session_state.data[selected_year]['fee_split']['mgmt_dao'] = st.session_state[dao_mgmt_key] / 100
                    st.session_state.data[selected_year]['fee_split']['mgmt_operator'] = (100 - st.session_state[dao_mgmt_key]) / 100

            st.session_state.has_changes = False
            st.session_state.data_modified = True  # Mark that data differs from original
            st.success("✅ All changes applied!")
            st.rerun()

with col_reset:
    st.markdown("<div style='margin-top: 1.7rem;'></div>", unsafe_allow_html=True)
    if st.button("🔄 Reset to Original Scenario", use_container_width=True, disabled=not st.session_state.data_modified):
        # Reload from Excel
        excel_file = 'Makina Revenue Generation Estimates.xlsx'
        st.session_state.data = load_excel_data(excel_file)
        st.session_state.has_changes = False
        st.session_state.data_modified = False  # Back to original
        st.success("✅ Reset to original scenario!")
        st.rerun()

st.markdown("---")

# Get data for selected year
year_data = st.session_state.data[selected_year]

# Store current slider values to detect changes later
if 'slider_values' not in st.session_state:
    st.session_state.slider_values = {}

# Calculate aggregate metrics
total_dao_revenue = 0
total_tvl = 0
for asset in ['usdc', 'eth', 'btc']:
    metrics = calculate_asset_metrics(year_data[asset], year_data['fee_split'])
    total_dao_revenue += metrics['dao_revenue']
    total_tvl += metrics['tvl_usd']

avg_dao_take_rate = (total_dao_revenue / total_tvl * 100) if total_tvl > 0 else 0

# Top 4 sections side-by-side
section1, section2, section3, section4 = st.columns(4)

# Section 1: TVL (big number)
with section1:
    st.markdown("<h2 style='margin-bottom: 0.5rem;'>TVL</h2>", unsafe_allow_html=True)
    # Extra large TVL metric taking space of two regular metrics
    st.markdown(f"<h1 style='font-size: 3.5rem; margin-top: 0; margin-bottom: 0;'>{format_number(total_tvl)}</h1>", unsafe_allow_html=True)

# Section 2: DAO Revenue & Take Rate
with section2:
    st.markdown("### DAO Revenue &<br>Take Rate", unsafe_allow_html=True)
    st.metric("Avg. DAO Take Rate", f"{avg_dao_take_rate:.2f}%")
    # Add spacing before second metric to align with sections 3 and 4
    st.markdown("<div style='margin-top: 0.2rem;'></div>", unsafe_allow_html=True)
    st.metric("Annualized DAO Revenue", format_number(total_dao_revenue))

# Section 3: Projected Revenue for Buyback and Token Staking Rewards
with section3:
    st.markdown("### Buyback & Staking<br>Rewards", unsafe_allow_html=True)

    # Slider for buyback percentage
    buyback_pct = st.slider(
        "% of Revenue",
        min_value=40.0,
        max_value=80.0,
        value=year_data['buyback_pct'],
        step=1.0,
        format="%.0f%%",
        key=f'buyback_slider_{selected_year}'
    )

    # Calculate buyback revenue
    buyback_revenue = total_dao_revenue * (buyback_pct / 100)

    st.metric("Revenue for Buyback & Staking", format_number(buyback_revenue))

# Section 4: Projected FDV
with section4:
    st.markdown("### Projected FDV<br>&nbsp;", unsafe_allow_html=True)  # &nbsp; for empty second line

    # P/Rev slider
    price_rev = st.slider(
        "P/Rev",
        min_value=15.0,
        max_value=45.0,
        value=year_data['valuation']['price_rev_ratio'],
        step=1.0,
        format="%.0f",
        key=f'price_rev_slider_{selected_year}'
    )

    # Calculate FDV
    fdv = total_dao_revenue * price_rev

    st.metric("FDV", format_number(fdv))

st.markdown("---")

# Check for unsaved changes
has_unsaved_changes = False

# Check if buyback_pct or price_rev changed
if (abs(buyback_pct - year_data['buyback_pct']) > 0.01 or
    abs(price_rev - year_data['valuation']['price_rev_ratio']) > 0.01):
    has_unsaved_changes = True

# Asset sections (USDC, ETH, BTC)
for asset_key, asset_name in [('usdc', 'USDC'), ('eth', 'ETH'), ('btc', 'BTC')]:
    st.subheader(f"{asset_name} Vault")

    asset_data = year_data[asset_key]

    # Create three columns: Inputs, Flow Chart, Outputs
    col_input, col_chart, col_output = st.columns([1, 2, 1])

    with col_input:
        st.markdown("**Inputs**")

        # APR - Slider
        apr_input = st.slider(
            f"APR (%)",
            min_value=0.0,
            max_value=20.0,
            value=min(asset_data['performance_growth'] * 100, 20.0),
            step=0.1,
            format="%.1f",
            key=f'{asset_key}_apr_{selected_year}'
        )

        # TVL - Slider with asset-specific ranges
        if asset_key == 'usdc':
            # USDC: 0 to 10B
            max_tvl = 10_000_000_000
            step_tvl = 100_000_000  # 100M steps
            tvl_input = st.slider(
                f"TVL (USDC)",
                min_value=0.0,
                max_value=float(max_tvl),
                value=min(float(asset_data['tvl_units']), float(max_tvl)),
                step=float(step_tvl),
                format="%.0f",
                key=f'{asset_key}_tvl_{selected_year}'
            )
            st.caption(f"Current: {format_number(tvl_input)}")
        elif asset_key == 'eth':
            # ETH: 0 to 1M
            max_tvl = 1_000_000
            step_tvl = 10_000  # 10K steps
            tvl_input = st.slider(
                f"TVL (ETH)",
                min_value=0.0,
                max_value=float(max_tvl),
                value=min(float(asset_data['tvl_units']), float(max_tvl)),
                step=float(step_tvl),
                format="%.0f",
                key=f'{asset_key}_tvl_{selected_year}'
            )
            st.caption(f"Current: {tvl_input:,.0f} ETH")
        else:  # BTC
            # BTC: 0 to 50K
            max_tvl = 50_000
            step_tvl = 500  # 500 steps
            tvl_input = st.slider(
                f"TVL (BTC)",
                min_value=0.0,
                max_value=float(max_tvl),
                value=min(float(asset_data['tvl_units']), float(max_tvl)),
                step=float(step_tvl),
                format="%.0f",
                key=f'{asset_key}_tvl_{selected_year}'
            )
            st.caption(f"Current: {tvl_input:,.0f} BTC")

        # Price (only for ETH and BTC) - Keep as number input
        if asset_key != 'usdc':
            price_input = st.number_input(
                "Price (USD)",
                value=asset_data['price'],
                min_value=0.0,
                step=100.0 if asset_key == 'eth' else 1000.0,
                format="%.0f",
                key=f'{asset_key}_price_{selected_year}'
            )
        else:
            price_input = 1.0

        st.markdown("**Fee Structure**")

        # Performance fee - Slider
        perf_fee_input = st.slider(
            "Performance Fee (%)",
            min_value=0.0,
            max_value=20.0,
            value=min(asset_data['perf_fee'] * 100, 20.0),
            step=0.1,
            format="%.1f",
            key=f'{asset_key}_perf_fee_{selected_year}'
        )

        # Management fee - Slider
        mgmt_fee_input = st.slider(
            "Management Fee (%)",
            min_value=0.0,
            max_value=3.0,
            value=min(asset_data['mgmt_fee'] * 100, 3.0),
            step=0.01,
            format="%.2f",
            key=f'{asset_key}_mgmt_fee_{selected_year}'
        )

        # DAO share in Performance fee - Slider
        dao_perf_input = st.slider(
            "DAO Share - Perf Fee (%)",
            min_value=0.0,
            max_value=100.0,
            value=year_data['fee_split']['perf_dao'] * 100,
            step=1.0,
            format="%.0f",
            key=f'{asset_key}_dao_perf_{selected_year}'
        )

        # DAO share in Management fee - Slider
        dao_mgmt_input = st.slider(
            "DAO Share - Mgmt Fee (%)",
            min_value=0.0,
            max_value=100.0,
            value=year_data['fee_split']['mgmt_dao'] * 100,
            step=1.0,
            format="%.0f",
            key=f'{asset_key}_dao_mgmt_{selected_year}'
        )

        # Check if any value has changed from saved data
        if (abs(apr_input / 100 - asset_data['performance_growth']) > 0.001 or
            abs(tvl_input - asset_data['tvl_units']) > 0.01 or
            abs(price_input - asset_data['price']) > 0.01 or
            abs(perf_fee_input / 100 - asset_data['perf_fee']) > 0.001 or
            abs(mgmt_fee_input / 100 - asset_data['mgmt_fee']) > 0.0001 or
            abs(dao_perf_input / 100 - year_data['fee_split']['perf_dao']) > 0.01 or
            abs(dao_mgmt_input / 100 - year_data['fee_split']['mgmt_dao']) > 0.01):
            has_unsaved_changes = True

        # Update button
        if st.button(f"Update {asset_name}", key=f'update_{asset_key}_{selected_year}'):
            st.session_state.data[selected_year][asset_key]['performance_growth'] = apr_input / 100
            st.session_state.data[selected_year][asset_key]['tvl_units'] = tvl_input
            st.session_state.data[selected_year][asset_key]['price'] = price_input
            st.session_state.data[selected_year][asset_key]['perf_fee'] = perf_fee_input / 100
            st.session_state.data[selected_year][asset_key]['mgmt_fee'] = mgmt_fee_input / 100
            st.session_state.data[selected_year]['fee_split']['perf_dao'] = dao_perf_input / 100
            st.session_state.data[selected_year]['fee_split']['mgmt_dao'] = dao_mgmt_input / 100
            st.session_state.data[selected_year]['fee_split']['perf_operator'] = (100 - dao_perf_input) / 100
            st.session_state.data[selected_year]['fee_split']['mgmt_operator'] = (100 - dao_mgmt_input) / 100
            st.session_state.has_changes = False
            st.session_state.data_modified = True  # Mark that data differs from original
            st.rerun()

    # Calculate current metrics using current input values
    current_data = {
        'tvl_units': tvl_input,
        'price': price_input,
        'mgmt_fee': mgmt_fee_input / 100,
        'perf_fee': perf_fee_input / 100,
        'performance_growth': apr_input / 100,
    }

    current_fee_split = {
        'mgmt_dao': dao_mgmt_input / 100,
        'mgmt_operator': (100 - dao_mgmt_input) / 100,
        'perf_dao': dao_perf_input / 100,
        'perf_operator': (100 - dao_perf_input) / 100,
    }

    metrics = calculate_asset_metrics(current_data, current_fee_split)

    with col_chart:
        # Flow chart with actual fee splits
        fig = create_flow_chart(
            metrics,
            asset_name,
            dao_perf_share=dao_perf_input / 100,
            dao_mgmt_share=dao_mgmt_input / 100
        )
        st.plotly_chart(fig, use_container_width=True)

    with col_output:
        st.markdown("<div style='margin-left: 30px;'>", unsafe_allow_html=True)
        st.markdown("**Outputs**")
        st.markdown("")  # Add spacing

        # Row 1: TVL / LP net return
        row1_col1, row1_col2 = st.columns(2)
        with row1_col1:
            st.metric(
                label="TVL (USD)",
                value=format_number_short(metrics['tvl_usd']),
                label_visibility="visible"
            )
        with row1_col2:
            st.metric(
                label="LP Net Return",
                value=f"{metrics['lp_apr']:.2f}%",
                label_visibility="visible"
            )

        st.markdown("")  # Add spacing

        # Row 2: DAO take rate / Annualized DAO rev
        row2_col1, row2_col2 = st.columns(2)
        with row2_col1:
            st.metric(
                label="DAO Take Rate",
                value=f"{metrics['dao_take_rate']:.2f}%",
                label_visibility="visible"
            )
        with row2_col2:
            st.metric(
                label="Ann. DAO Rev",
                value=format_number_short(metrics['dao_revenue']),
                label_visibility="visible"
            )

        st.markdown("")  # Add spacing

        # Row 3: Operator take rate / Ann. Operator revenue
        row3_col1, row3_col2 = st.columns(2)
        with row3_col1:
            st.metric(
                label="Operator Take Rate",
                value=f"{metrics['op_take_rate']:.2f}%",
                label_visibility="visible"
            )
        with row3_col2:
            st.metric(
                label="Ann. Op Rev",
                value=format_number_short(metrics['op_revenue']),
                label_visibility="visible"
            )

        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("---")

# Update has_changes flag based on detected unsaved changes
if has_unsaved_changes:
    st.session_state.has_changes = True
elif not has_unsaved_changes and st.session_state.has_changes:
    # Keep the flag if it was set (until Apply or Reset is clicked)
    pass
